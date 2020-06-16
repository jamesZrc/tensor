import re
from collections import Counter
import openpyxl

excel_path = input("excel路径: ")
sheet_index = input("sheet序号: ")
wc_col = input("统计列: ")

# excel_path = "C:/Users/james/Desktop/a.xlsx"
# sheet_index = "1"
# wc_col = "1"

book = openpyxl.load_workbook(excel_path)
sheet = book.worksheets[int(sheet_index) - 1]
speech = []
for row in range(sheet.max_row):
    value = re.sub(r'!|@|#|$|%|^|&|\*|\(|\)|[|]|{|}|"|,|\?|\.|\'|-|_|\d|:|？|；|，|。', " ",
                   sheet.cell(row + 1, int(wc_col)).value.lower())
    speech.extend(value.split())

wd = Counter(speech)

wc_sheet = book.create_sheet('词频统计')
wc_sheet.cell(1, 1, "单词")
wc_sheet.cell(1, 2, "频率")

item_size = wd.items()
i = 2
for key, value in sorted(wd.items(), key=lambda kv: (kv[0], kv[1])):
    wc_sheet.cell(i, 1, key)
    wc_sheet.cell(i, 2, value)
    i = i + 1

book.save(excel_path)
input("统计完成，按enter退出！")
